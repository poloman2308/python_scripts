from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook

from combined_summaries import create_combined_summaries
from npl_hw_eol_timeline import create_npl_hw_eol_timeline
from npl_sw_eol_timeline import create_npl_sw_eol_timeline

file_path = ""

def chassis_repot_enriching(file_path):
    try:
        wb = load_workbook(file_path)
    except FileNotFoundError:
        wb = Workbook()

    check_headers(wb)
    create_npl_hw_eol_timeline(wb)
    create_npl_sw_eol_timeline(wb)
    create_combined_summaries(wb)

    wb.save(file_path)
    
    