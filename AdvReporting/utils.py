import re
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.chart import BarChart
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

standart_bold_font = Font(color='FFFFFFFF', bold=True)
standart_pattern_fill = PatternFill(
    start_color='00008B',
    end_color='00008B',
    fill_type='solid'
)

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def ensure_tabs_exist(work_book: Workbook, tab_names):
    existing_tabs = work_book.sheetnames
    # If any tab is missing, create it
    for tab_name in tab_names:
        if tab_name not in existing_tabs:
            work_book.create_sheet(tab_name)


def get_end_of_quarter(date):
    quarter = (date.month - 1) // 3 + 1
    if quarter == 1:
        return datetime(date.year, 3, 31)
    elif quarter == 2:
        return datetime(date.year, 6, 30)
    elif quarter == 3:
        return datetime(date.year, 9, 30)
    else:
        return datetime(date.year, 12, 31)
    

def input_dates(work_sheet, start_cell, end_cell):
    
    # Get current date and calculate the end of the current quarter
    current_date = datetime.now()
    end_of_quarter_date = get_end_of_quarter(current_date)
    work_sheet[start_cell] = end_of_quarter_date
    
    # Calculate the end of the next quarters and input into subsequent cells
    start_cell_obj = work_sheet[start_cell]
    end_cell_obj = work_sheet[end_cell]
    current_col, row = start_cell_obj.column, start_cell_obj.row
    end_col = end_cell_obj.column
    
    while current_col < end_col:
        # Move to the start of the next quarter
        if end_of_quarter_date.month == 3:
            next_date = datetime(end_of_quarter_date.year, 4, 1)
        elif end_of_quarter_date.month == 6:
            next_date = datetime(end_of_quarter_date.year, 7, 1)
        elif end_of_quarter_date.month == 9:
            next_date = datetime(end_of_quarter_date.year, 10, 1)
        else:
            next_date = datetime(end_of_quarter_date.year + 1, 1, 1)
        
        end_of_quarter_date = get_end_of_quarter(next_date)
        work_sheet.cell(row=row, column=current_col + 1, value=end_of_quarter_date)
        current_col += 1


def input_values(work_sheet, cell_values, formulas=None):
    if formulas is None:
        formulas = {}
    for cell, value in cell_values.items():
        work_sheet[cell] = value
    for cell_range, formula in formulas.items():
        for row in work_sheet[cell_range]:
            for cell in row:
                cell.value = formula


def calculate_values(work_sheet, calculations):
    for cell, formula in calculations.items():
        work_sheet[cell] = formula


def create_bar_chart(
    data_categories,
    data_values,
    title,
    x_axis_title=None,
    y_axis_title=None,
    bar_color="FF0000",
    chart_width=25,
    legend_pos_b=False,
):
    # Create a new clustered column chart
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = title
    chart.x_axis.title = x_axis_title
    chart.y_axis.title = y_axis_title
    chart.add_data(data_values, titles_from_data=True)
    chart.set_categories(data_categories)

    # Customize the color of the bars to red
    s = chart.series[0]
    s.graphicalProperties.solidFill = bar_color

    chart.width = chart_width
    if legend_pos_b:
        chart.legend.legendPos = 'b'
    else:
        # Remove the legend
        chart.legend = None

    return chart


def format_cells(work_sheet, format):
    for cell_range, cell_format in format.items():
        cells = work_sheet[cell_range]
        font = cell_format.get('font', None)
        fill = cell_format.get('fill', None)
        number_format = cell_format.get('number_format', None)
        wrap_text = cell_format.get('wrap_text', None)
        align_right = cell_format.get('align_right', None)
        border = cell_format.get('border', None)

        # Apply styles to each cell in the range
        if ":" in cell_range:
            for row in cells:
                for cell in row:
                    if font is not None:
                        cell.font = font
                    if fill is not None:
                        cell.fill = fill
                    if number_format is not None:
                        cell.number_format = number_format
                    if wrap_text is not None:
                        cell.alignment = Alignment(wrap_text=wrap_text)
                    if align_right is not None:
                        cell.alignment = Alignment(horizontal='right')
                    if border is not None:
                        cell.border = border
        else:
            cell = cells
            if font is not None:
                cell.font = font
            if fill is not None:
                cell.fill = fill
            if number_format is not None:
                cell.number_format = number_format
            if wrap_text is not None:
                cell.alignment = Alignment(wrap_text=wrap_text)
            if align_right is not None:
                cell.alignment = Alignment(horizontal='right')
            if border is not None:
                cell.border = border

                
def column_sizing(work_sheet):
    sheet_name = work_sheet.title

    if sheet_name == 'NPL - HW EOL Timeline':
        work_sheet.column_dimensions['B'].width = 32

        work_sheet.column_dimensions['AA'].width = 12

        for col in range(ord('F'), ord('Y') + 1):
            work_sheet.column_dimensions[chr(col)].width = 12  

        for col in range(ord('AB'), ord('AF') + 1):
            col_letter = chr(col - ord('A') + ord('A')) if col <= ord('Z') else 'A' + chr(col - ord('Z') - 1)
            work_sheet.column_dimensions[col_letter].width = 16

    elif sheet_name == 'NPL - SW EOL Timeline':
        work_sheet.column_dimensions['B'].width = 32

        for col in range(ord('F'), ord('Y') + 1):
            work_sheet.column_dimensions[chr(col)].width = 12

    elif sheet_name == 'Combined Summaries':
        work_sheet.column_dimensions['A'].width = 32

        for col in range(ord('C'), ord('F') + 1):
            work_sheet.column_dimensions[chr(col)].width = 16

            
def hide_columns(work_sheet, columns):
    for col in columns:
        work_sheet.column_dimensions[get_column_letter(col)].hidden = True

        
def check_headers(workbook):
    """
    Comapre existing headers to the expected headers and rise exception
    if they don't match.
    Args:
        workbook:
    Returns:
    """
    ws = workbook['Report']
    headers_expected = [
        'serial_number',
        'source',
        'jpmc_product_owner',
        'jpmc_nash_role',
        'cisco_cx_technology_team',
        'host_name',
        'orderable_pid',
        "cons_ord_pid",
        'base_pid',
        'device_type',
        'shorten_product_family',
        'hw_first_customer_ship',
        'hw_end_of_life_announcement',
        'hw_end_of_sale',
        'end_of_failure_analysis',
        'hw_last_date_of_support',
        'hw_current_milestone',
        'hw_bulletin_number',
        "sw_version",
        "sw_type",
        "sw_version_source",
        'sw_end_of_life_announcement',
        'sw_end_of_sale',
        'end_of_sw_maintenance',
        'sw_end_of_vulnerability_support',
        'sw_last_date_of_support',
        'sw_current_milestone',
        'sw_bulletin_number',
        'location_id',
        'building',
        'street_address',
        'city',
        'state',
        'postal_code',
        'country',
        'region',
        'dc_grouping',
    ]
    headers_actual = [cell.value for cell in ws[1]]
    if headers_actual != headers_expected:
        raise ValueError(f"Headers do not match. Expected: {headers_expected}, Actual: {headers_actual}")

