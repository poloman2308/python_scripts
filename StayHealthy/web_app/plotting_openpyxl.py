import matplotlib.ticker as ticker
import matplotlib.pyplot as plt
from io import BytesIO
import numpy as np
from openpyxl.drawing.image import Image
from openpyxl.drawing import drawing
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    NamedStyle,
    PatternFill,
    Side,
)

WIDTH = 1520
plt.switch_backend('Agg')

def merge_cells(cell, width, height):
    cell.parent.merge_cells(
        start_row=cell.row,
        start_column=cell.column,
        end_row=cell.row+height-1,
        end_column=cell.column+width-1,
    )

def get_color(color):
    colors = {
        "red": "#dc4c43",
        "orange": "#e68c4f",
        "yellow": "#fbbc34",
        "green": "#54a353",
        "blue": "#00b0f0",
        "hw_green_1": "#53a050",
        "hw_green_2": "#8ab453",
        "hw_green_3": "#c0c856",
        "hw_yellow": "#f7dc58",
        "hw_orange_1": "#f1b552",
        "hw_orange_2": "#e88e4a",
        "hw_orange_3": "#e16743",
        "hw_red": "#d93f3c",
        "grey": "#292b2c",
        "white": "#FFFFFF"
    }
    return colors[color]

default_border = Border(
    left=Side(border_style='thin', color='FF000000'),
    right=Side(border_style='thin', color='FF000000'),
    top=Side(border_style='thin', color='FF000000'),
    bottom=Side(border_style='thin', color='FF000000'),
)

def get_visual(width, height, font_color, size, bold, bg, v, h):
    # half_options = {
    return {
        'width': (WIDTH*width)-(1/width),
        'height': height,
        'font': {'color': font_color,
                'size': size,
                'bold': bold},
        'align': {'vertical': v,
                'horizontal': h
                },
        'gradient': {'colors': [bg, bg]},
    }


def get_style(
    style_name: str,
    font_color: str,
    size: int,
    bold: bool,
    italic: bool,
    underline: bool,
    bg_color: str,
    vertical_alignment: str,
    horizontal_alignment: str,
    border: bool = False,
):
    if underline:
        underline = 'single'
    else:
        underline = None
    style = NamedStyle(name=style_name)
    style.font = Font(
        size=size,
        bold=bold,
        italic=italic,
        underline=underline,
        color=font_color,
    )
    style.alignment = Alignment(
        horizontal=horizontal_alignment,
        vertical=vertical_alignment,
        wrap_text=True,
    )
    style.fill = PatternFill(
        fill_type="solid",
        start_color=bg_color,
        end_color=bg_color,
    )
    if border:
        style.border = default_border
    return style

style_options = get_style(
    style_name="style_options",
    font_color=get_color("white")[1:],
    size=22,
    bold=True,
    italic=False,
    underline=False,
    bg_color=get_color("grey")[1:],
    vertical_alignment='center',
    horizontal_alignment='center',
    border=True,
)
style_standard_options = get_style(
    style_name="style_standard_options",
    font_color=get_color("white")[1:],
    size=20,
    bold=True,
    italic=False,
    underline=False,
    bg_color=get_color("blue")[1:],
    vertical_alignment='center',
    horizontal_alignment='center',
    border=True,
)
style_options_green = get_style(
    style_name="style_options_green",
    font_color=get_color("white")[1:],
    size=22,
    bold=True,
    italic=False,
    underline=False,
    bg_color=get_color("green")[1:],
    vertical_alignment='center',
    horizontal_alignment='center',
    border=True,
)
style_options_yellow = get_style(
    style_name="style_options_yellow",
    font_color=get_color("white")[1:],
    size=22,
    bold=True,
    italic=False,
    underline=False,
    bg_color=get_color("yellow")[1:],
    vertical_alignment='center',
    horizontal_alignment='center',
    border=True,
)
style_options_red = get_style(
    style_name="style_options_red",
    font_color=get_color("white")[1:],
    size=22,
    bold=True,
    italic=False,
    underline=False,
    bg_color=get_color("red")[1:],
    vertical_alignment='center',
    horizontal_alignment='center',
    border=True,
)

style_big_options = get_style(
    style_name="style_big_options",
    font_color=get_color("grey")[1:],
    size=28,
    bold=True,
    italic=False,
    underline=False,
    bg_color=get_color("white")[1:],
    vertical_alignment='center',
    horizontal_alignment='center',
    border=True,
)
style_standard_big_options = get_style(
    style_name="style_standard_big_options",
    font_color=get_color("white")[1:],
    size=28,
    bold=True,
    italic=False,
    underline=False,
    bg_color=get_color("blue")[1:],
    vertical_alignment='center',
    horizontal_alignment='center',
    border=True,
)
style_big_options_green = get_style(
    style_name="style_big_options_green",
    font_color=get_color("white")[1:],
    size=36,
    bold=True,
    italic=False,
    underline=False,
    bg_color=get_color("green")[1:],
    vertical_alignment='center',
    horizontal_alignment='center',
    border=True,
)
style_big_options_yellow = get_style(
    style_name="style_big_options_yellow",
    font_color=get_color("white")[1:],
    size=36,
    bold=True,
    italic=False,
    underline=False,
    bg_color=get_color("yellow")[1:],
    vertical_alignment='center',
    horizontal_alignment='center',
    border=True,
)
style_big_options_orange = get_style(
    style_name="style_big_options_orange",
    font_color=get_color("white")[1:],
    size=36,
    bold=True,
    italic=False,
    underline=False,
    bg_color=get_color("orange")[1:],
    vertical_alignment='center',
    horizontal_alignment='center',
    border=True,
)
style_big_options_red = get_style(
    style_name="style_big_options_red",
    font_color=get_color("white")[1:],
    size=36,
    bold=True,
    italic=False,
    underline=False,
    bg_color=get_color("red")[1:],
    vertical_alignment='center',
    horizontal_alignment='center',
    border=True,
)
style_small_options_txt = get_style(
    style_name="style_small_options_txt",
    font_color=get_color("grey")[1:],
    size=14,
    bold=True,
    italic=False,
    underline=False,
    bg_color=get_color("white")[1:],
    vertical_alignment='center',
    horizontal_alignment='left',
    border=True,
)
        
def nth_repl_all(s, sub, repl, nth):
    find = s.find(sub)
    # loop util we find no match
    i = 1
    while find != -1:
        # if i  is equal to nth we found nth matches so replace
        if i == nth:
            s = s[:find]+repl+s[find + len(sub):]
            i = 0
        # find + len(sub) + 1 means we start after the last match
        find = s.find(sub, find + len(sub) + 1)
        i += 1
    return s

def create_pie(worksheet, dataframe, plot_type, color, title, x_label, y_label, x_position, y_position, bar_label=True, figsize_x=6.2, figsize_y=4.15, summary=None, rotation=45, ha='right'):
    bbox_props=dict(boxstyle='square,pad=0.3',fc ='w',ec='k',lw=1.72)
    kw=dict(xycoords='data',textcoords='data',arrowprops=dict(arrowstyle='-'),zorder=0,va='center')

    dataframe = dataframe.sort_index()
    labels = dataframe.index.tolist()
    values = [item for sublist in dataframe.values.tolist() for item in sublist]
    colors=['#b2c24f','#a04950', '#61aba8', '#274465', 
            '#687c8a', '#ebcb47', '#b38d67', '#4d3a6a', '#789f70', '#667c89']
    fig1,ax1=plt.subplots(figsize=(figsize_x,figsize_y))
    annotate_dict = {k:v for k,v in zip(labels, values)}
    val = [[x,y] for x,y in zip(sorted(values, reverse=True),sorted(values))]
    values1 = sum(val, [])

    if not summary:
        colors=['#b2c24f','#a04950', '#61aba8', '#274465', 
                '#687c8a', '#ebcb47', '#b38d67', '#4d3a6a', '#789f70', '#667c89']
        wedges,texts=ax1.pie(values1[:len(values)],explode=[0.01] * len(labels),textprops={"fontsize":8}, labeldistance=2.5,startangle=50, colors=colors)
    else:
        sorted_labels = []
        for cnt in values1[:len(values)]:
            for label in labels:
                if annotate_dict[label] == cnt:
                    sorted_labels.append(label)
        labels = sorted_labels
        colors={'Add Coverage': get_color('orange'), 'No Action Required': get_color('green'),
                'Refresh Hardware': get_color('red'), 'Upgrade Software': get_color('yellow')}
        wedges,texts=ax1.pie(values1[:len(values)],explode=[0.01] * len(labels),textprops={"fontsize":10},labeldistance=1.8,startangle=50, colors=[colors[key] for key in labels])

    new_labels = []
    for v in values1[:len(values)]:
        for key, value in annotate_dict.items():
            if (int(v) == int(value)):
                if len(key) > 35:
                    new_label = nth_repl_all(key, " ", "\n", 4)
                    if new_label not in new_labels:
                        new_labels.append(new_label)
                        break
                elif len(key) < 30:
                    new_label = key
                    if new_label not in new_labels:
                        new_labels.append(new_label)
                        break
                else:
                    new_label = nth_repl_all(key, " ", "\n", 3)
                    if new_label not in new_labels:
                        new_labels.append(new_label)
                        break
            
    for i,p in enumerate(wedges):
        ang=(p.theta2-p.theta1)/2. +p.theta1
        y=np.sin(np.deg2rad(ang))
        x=np.cos(np.deg2rad(ang))
        horizontalalignment={-1:"right",1:"left"}[int(np.sign(x))]
        connectionstyle="angle,angleA=0,angleB={}".format(ang)
        kw["arrowprops"].update({"connectionstyle":connectionstyle})
        if not summary:
            ax1.annotate(new_labels[i],xy=(x, y),xytext=(1.35*np.sign(x),1.8*y),
                        horizontalalignment=horizontalalignment,**kw)
        else:
            ax1.annotate(new_labels[i],xy=(x, y),xytext=(1.35*np.sign(x),1.4*y),
                        horizontalalignment=horizontalalignment,**kw)
    fig1 = plt.gcf()
    plt.title(title)
    imgdata = BytesIO()
    plt.savefig(imgdata, format="png", bbox_inches='tight')
    imgdata.seek(0)
    img = Image(imgdata)
    img.anchor = worksheet.cell(y_position, x_position).coordinate
    worksheet.add_image(img)

def create_plot(
    worksheet,
    dataframe,
    plot_type,
    color,
    title,
    x_label,
    y_label,
    x_position,
    y_position,
    bar_label=True,
    figsize_x=None,
    figsize_y=None,
    summary=None,
    rotation=45,
    ha='right',
    legend=True,
    percentage=False,
):
    if plot_type == 'bar':
        if not figsize_x: figsize_x=5.9
        if not figsize_y: figsize_y=3
        ax = dataframe.plot(kind=plot_type, color=color, figsize=(figsize_x, figsize_y))
        # ax.bar_label(ax.containers[0])
        plt.xticks(rotation=rotation, ha=ha)
        plt.xlabel(x_label)
        plt.ylabel(y_label)
        plt.margins(0.05, 0.1)
        plt.title(title)
        if percentage:
            ax.yaxis.set_major_formatter(ticker.PercentFormatter(xmax=1))
    if plot_type == 'barh':
        if not figsize_x: figsize_x=3.5
        if not figsize_y: figsize_y=3.2
        ax = dataframe.plot(kind=plot_type, color=color, figsize=(figsize_x, figsize_y))
        ax.bar_label(ax.containers[0])
        # plt.xticks(rotation=70)
        plt.xlabel(x_label)
        plt.ylabel(y_label) 
        plt.margins(0.2, 0.1)
        plt.title(title)
        if percentage:
            ax.xaxis.set_major_formatter(ticker.PercentFormatter())
    if plot_type == 'pie':
        if not figsize_x: figsize_x=6.2
        if not figsize_y: figsize_y=4.15
        if not summary:
            dataframe.plot(kind=plot_type, colors=['#b2c24f','#a04950', '#61aba8', '#274465', 
                        '#687c8a', '#ebcb47', '#b38d67', '#4d3a6a', '#789f70', '#667c89'], 
                        subplots=True, figsize=(figsize_x, figsize_y), legend=None, ylabel="", startangle=90)
        else:
            dataframe.plot(kind=plot_type, colors=[get_color('orange'), get_color('green'), get_color('red'), get_color('yellow')],
                        subplots=True, figsize=(figsize_x, figsize_y), legend=None, ylabel="", startangle=90)
        plt.title(title)
    if not legend:
        plt.legend().remove()
    imgdata = BytesIO()
    # plt.tight_layout()
    plt.savefig(imgdata, format="png", bbox_inches='tight')
    imgdata.seek(0)
    img = Image(imgdata)
    img.anchor = worksheet.cell(y_position, x_position).coordinate
    worksheet.add_image(img)