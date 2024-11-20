from datetime import date

import pandas as pd
from openpyxl.worksheet.worksheet import Worksheet

from utils.web_app import \
    plotting_openpyxl as plotting
from openpyxl.workbook import Workbook


class SWStrategy:
    def __init__(self, workbook: Workbook, devices):
        self.workbook = workbook
        self.devices = devices
        self.devices = self.devices.loc[((self.devices['SW Type'] != 'Not Applicable') & (self.devices['SW Type'] != 'Not available'))]
        self.sw_ldos_this_year = self.devices.loc[(self.devices['Update Software This Year'] == 'YES')]
        self.sw_ldos_next_year = self.devices.loc[(self.devices['Update Software Next Year'] == 'YES')]
        self.sw_ldos_in_2_years = self.devices.loc[(self.devices['Update Software In 2 Years'] == 'YES')]

    def create(self):
        worksheet: Worksheet = self.workbook.create_sheet('SW Strategy')

        worksheet.column_dimensions["A"].width = 2.6
        worksheet.column_dimensions["J"].width = 2.6
        worksheet.column_dimensions["S"].width = 2.6
        worksheet.column_dimensions["AB"].width = 2.6

        self.stats_this_year(worksheet)
        self.stats_next_year(worksheet)
        self.stats_in_2_years(worksheet)

        worksheet['B74'] = f'SW Diversity by Product Family'
        worksheet['B74'].style = plotting.style_standard_big_options
        plotting.merge_cells(
            worksheet['B74'],
            26,
            3
        )

        sw_diversity = self.devices.pivot_table(
            index=['Shortened Product Family'],
            values=['SW Version'],
            aggfunc=lambda x: len(x.unique()),
        ).sort_values(
            ('SW Version'),
            ascending=False
        ).head(20)

        sw_diversity = sw_diversity.rename(
            columns={
                "SW Version": "SW Versions",
                'Shortened Product Family': 'Product Family'
            }
        )

        plotting.create_plot(
            worksheet,
            sw_diversity,
            'bar',
            '#0070c0',
            None,
            'Product Family',
            'SW Versions',
            2,
            78,
            True,
            21.5,
            6,
            legend=False,
        )

        worksheet.sheet_view.showGridLines = False

    def stats_this_year(self, worksheet):
        worksheet['B2'] = f'Devices EoSWM by CX Tech Team through {date.today().year}'
        worksheet['B2'].style = plotting.style_standard_options
        plotting.merge_cells(
            worksheet['B2'],
            8,
            3
        )
        ldos_top_cx_tech_team_this = self.sw_ldos_this_year.pivot_table(
            index=['Cisco CX Technology Team'],
            values=['Serial Number'],
            aggfunc='count',
        )
        ldos_top_cx_tech_team_general = self.devices.pivot_table(
            index=['Cisco CX Technology Team'],
            values=['Serial Number'],
            aggfunc='count',
        )
        ldos_top_cx_tech_team_this = ldos_top_cx_tech_team_this.rename(
            columns={"Serial Number": "Components1"}
        )
        ldos_top_cx_tech_team_general = ldos_top_cx_tech_team_general.rename(
            columns={"Serial Number": "Components2"}
        )
        ldos_top_cx_tech_team_this = pd.merge(
            ldos_top_cx_tech_team_this,
            ldos_top_cx_tech_team_general,
            how='left',
            on=['Cisco CX Technology Team'],
        )
        ldos_top_cx_tech_team_this['Components'] = round(ldos_top_cx_tech_team_this['Components1'] / ldos_top_cx_tech_team_this['Components2'],4)
        ldos_top_cx_tech_team_this.drop(columns=['Components1', 'Components2'], inplace=True)
        plotting.create_plot(
            worksheet,
            ldos_top_cx_tech_team_this,
            'bar',
            '#0070c0',
            None,
            'CX Technology Team',
            None,
            2,
            6,
            True,
            6.5,
            3.7,
            legend=False,
            percentage=True,
        )
        worksheet['B26'] = f'Top 10 EoSWM Product Families through {date.today().year}'
        worksheet['B26'].style = plotting.style_standard_options
        plotting.merge_cells(
            worksheet['B26'],
            8,
            3
        )
        ldos_top_10_this = self.sw_ldos_this_year.pivot_table(
            index=['Shortened Product Family'],
            values=['Serial Number'],
            aggfunc='count'
        ).sort_values(
            ('Serial Number'),
            ascending=True
        ).tail(10)
        ldos_top_10_this = ldos_top_10_this.rename(
            columns={
                "Serial Number": "Components",
                'Shortened Product Family': 'Product Family'
            }
        )
        plotting.create_plot(
            worksheet,
            ldos_top_10_this,
            'barh',
            '#0070c0',
            None,
            None,
            'Product Family',
            2,
            30,
            True,
            5.4,
            4.5,
            legend=False,
        )
        worksheet['B50'] = f'Top 20 EoSWM Orderable PIDs through {date.today().year}'
        worksheet['B50'].style = plotting.style_standard_options
        plotting.merge_cells(
            worksheet['B50'],
            8,
            3
        )
        ldos_top_20_this = self.sw_ldos_this_year.pivot_table(
            index=['Orderable PID (Cx-IBA)', 'SW Type', 'SW Version'],
            values=['Serial Number'],
            aggfunc='count'
        ).sort_values(
            ('Serial Number'),
            ascending=False
        ).head(20)
        ldos_top_20_this = ldos_top_20_this.rename(
            columns={
                "Serial Number": "Components",
                'Orderable PID (Cx-IBA)': 'Orderable PID',
            }
        )
        for row in range(0, len(ldos_top_20_this)):
            row_name = ldos_top_20_this.iloc[row].name
            worksheet[f'B{50 + 3 + row}'] = row_name[0]
            worksheet[f'B{50 + 3 + row}'].border = plotting.default_border
            plotting.merge_cells(
                worksheet[f'B{50 + 3 + row}'],
                3,
                1
            )
            worksheet[f'E{50 + 3 + row}'] = row_name[1]
            worksheet[f'E{50 + 3 + row}'].border = plotting.default_border
            plotting.merge_cells(
                worksheet[f'E{50 + 3 + row}'],
                2,
                1
            )
            worksheet[f'G{50 + 3 + row}'] = row_name[2]
            worksheet[f'G{50 + 3 + row}'].border = plotting.default_border
            plotting.merge_cells(
                worksheet[f'G{50 + 3 + row}'],
                2,
                1
            )
            worksheet[f'I{50 + 3 + row}'] = ldos_top_20_this.iloc[row].values[0]
            worksheet[f'I{50 + 3 + row}'].border = plotting.default_border

    def stats_next_year(self, worksheet):
        worksheet['K2'] = f'Devices EoSWM by CX Tech Team through {date.today().year+1}'
        worksheet['K2'].style = plotting.style_standard_options
        plotting.merge_cells(
            worksheet['K2'],
            8,
            3
        )
        ldos_top_cx_tech_team_next = self.sw_ldos_next_year.pivot_table(
            index=['Cisco CX Technology Team'],
            values=['Serial Number'],
            aggfunc='count'
        )
        ldos_top_cx_tech_team_general = self.devices.pivot_table(
            index=['Cisco CX Technology Team'],
            values=['Serial Number'],
            aggfunc='count',
        )
        ldos_top_cx_tech_team_next = ldos_top_cx_tech_team_next.rename(
            columns={"Serial Number": "Components1"}
        )
        ldos_top_cx_tech_team_general = ldos_top_cx_tech_team_general.rename(
            columns={"Serial Number": "Components2"}
        )
        ldos_top_cx_tech_team_next = pd.merge(
            ldos_top_cx_tech_team_next,
            ldos_top_cx_tech_team_general,
            how='left',
            on=['Cisco CX Technology Team'],
        )
        ldos_top_cx_tech_team_next['Components'] = round(ldos_top_cx_tech_team_next['Components1'] / ldos_top_cx_tech_team_next['Components2'],4)
        ldos_top_cx_tech_team_next.drop(columns=['Components1', 'Components2'], inplace=True)
        plotting.create_plot(
            worksheet,
            ldos_top_cx_tech_team_next,
            'bar',
            '#0070c0',
            None,
            'CX Technology Team',
            None,
            11,
            6,
            True,
            6.5,
            3.7,
            legend=False,
            percentage=True,
        )
        worksheet['K26'] = f'Top 10 EoSWM Product Families through {date.today().year+1}'
        worksheet['K26'].style = plotting.style_standard_options
        plotting.merge_cells(
            worksheet['K26'],
            8,
            3
        )
        ldos_top_10_next = self.sw_ldos_next_year.pivot_table(
            index=['Shortened Product Family'],
            values=['Serial Number'],
            aggfunc='count'
        ).sort_values(
            ('Serial Number'),
            ascending=True
        ).tail(10)
        ldos_top_10_next = ldos_top_10_next.rename(
            columns={
                "Serial Number": "Components",
                'Shortened Product Family': 'Product Family'
            }
        )
        plotting.create_plot(
            worksheet,
            ldos_top_10_next,
            'barh',
            '#0070c0',
            None,
            None,
            'Product Family',
            11,
            30,
            True,
            5.4,
            4.5,
            legend=False,
        )
        worksheet['K50'] = f'Top 20 EoSWM Orderable PIDs through {date.today().year+1}'
        worksheet['K50'].style = plotting.style_standard_options
        plotting.merge_cells(
            worksheet['K50'],
            8,
            3
        )
        ldos_top_20_next = self.sw_ldos_next_year.pivot_table(
            index=['Orderable PID (Cx-IBA)', 'SW Type', 'SW Version'],
            values=['Serial Number'],
            aggfunc='count'
        ).sort_values(
            ('Serial Number'),
            ascending=False
        ).head(20)
        ldos_top_20_next = ldos_top_20_next.rename(
            columns={
                "Serial Number": "Components",
                'Orderable PID (Cx-IBA)': 'Orderable PID'
            }
        )
        for row in range(0, len(ldos_top_20_next)):
            row_name = ldos_top_20_next.iloc[row].name
            worksheet[f'K{50 + 3 + row}'] = row_name[0]
            worksheet[f'K{50 + 3 + row}'].border = plotting.default_border
            plotting.merge_cells(
                worksheet[f'K{50 + 3 + row}'],
                3,
                1
            )
            worksheet[f'N{50 + 3 + row}'] = row_name[1]
            worksheet[f'N{50 + 3 + row}'].border = plotting.default_border
            plotting.merge_cells(
                worksheet[f'N{50 + 3 + row}'],
                2,
                1
            )
            worksheet[f'P{50 + 3 + row}'] = row_name[2]
            worksheet[f'P{50 + 3 + row}'].border = plotting.default_border
            plotting.merge_cells(
                worksheet[f'P{50 + 3 + row}'],
                2,
                1
            )
            worksheet[f'R{50 + 3 + row}'] = ldos_top_20_next.iloc[row].values[
                0]
            worksheet[f'R{50 + 3 + row}'].border = plotting.default_border


    def stats_in_2_years(self, worksheet):
        worksheet['T2'] = f'Devices EoSWM by CX Tech Team through {date.today().year+2}'
        worksheet['T2'].style = plotting.style_standard_options
        plotting.merge_cells(
            worksheet['T2'],
            8,
            3
        )
        ldos_top_cx_tech_team_in_2 = self.sw_ldos_in_2_years.pivot_table(
            index=['Cisco CX Technology Team'],
            values=['Serial Number'],
            aggfunc='count'
        )
        ldos_top_cx_tech_team_general = self.devices.pivot_table(
            index=['Cisco CX Technology Team'],
            values=['Serial Number'],
            aggfunc='count',
        )
        ldos_top_cx_tech_team_in_2 = ldos_top_cx_tech_team_in_2.rename(
            columns={"Serial Number": "Components1"}
        )
        ldos_top_cx_tech_team_general = ldos_top_cx_tech_team_general.rename(
            columns={"Serial Number": "Components2"}
        )
        ldos_top_cx_tech_team_in_2 = pd.merge(
            ldos_top_cx_tech_team_in_2,
            ldos_top_cx_tech_team_general,
            how='left',
            on=['Cisco CX Technology Team'],
        )
        ldos_top_cx_tech_team_in_2['Components'] = round(ldos_top_cx_tech_team_in_2['Components1'] / ldos_top_cx_tech_team_in_2['Components2'],4)
        ldos_top_cx_tech_team_in_2.drop(columns=['Components1', 'Components2'], inplace=True)
        plotting.create_plot(
            worksheet,
            ldos_top_cx_tech_team_in_2,
            'bar',
            '#0070c0',
            None,
            'CX Technology Team',
            None,
            20,
            6,
            True,
            6.5,
            3.7,
            legend=False,
            percentage=True,
        )
        worksheet['T26'] = f'Top 10 EoSWM Product Families through {date.today().year+2}'
        worksheet['T26'].style = plotting.style_standard_options
        plotting.merge_cells(
            worksheet['T26'],
            8,
            3
        )
        ldos_top_10_in_2 = self.sw_ldos_in_2_years.pivot_table(
            index=['Shortened Product Family'],
            values=['Serial Number'],
            aggfunc='count'
        ).sort_values(
            ('Serial Number'),
            ascending=True
        ).tail(10)
        ldos_top_10_in_2 = ldos_top_10_in_2.rename(
            columns={
                "Serial Number": "Components",
                'Shortened Product Family': 'Product Family'
            }
        )
        plotting.create_plot(
            worksheet,
            ldos_top_10_in_2,
            'barh',
            '#0070c0',
            None,
            None,
            'Product Family',
            20,
            30,
            True,
            5.4,
            4.5,
            legend=False,
        )
        worksheet['T50'] = f'Top 20 EoSWM Orderable PIDs through {date.today().year+2}'
        worksheet['T50'].style = plotting.style_standard_options
        plotting.merge_cells(
            worksheet['T50'],
            8,
            3
        )
        ldos_top_20_in_2 = self.sw_ldos_in_2_years.pivot_table(
            index=['Orderable PID (Cx-IBA)', 'SW Type', 'SW Version'],
            values=['Serial Number'],
            aggfunc='count'
        ).sort_values(
            ('Serial Number'),
            ascending=False
        ).head(20)
        ldos_top_20_in_2 = ldos_top_20_in_2.rename(
            columns={
                "Serial Number": "Components",
                'Orderable PID (Cx-IBA)': 'Orderable PID'
            }
        )
        for row in range(0, len(ldos_top_20_in_2)):
            row_name = ldos_top_20_in_2.iloc[row].name
            worksheet[f'T{50 + 3 + row}'] = row_name[0]
            worksheet[f'T{50 + 3 + row}'].border = plotting.default_border
            plotting.merge_cells(
                worksheet[f'T{50 + 3 + row}'],
                3,
                1
            )
            worksheet[f'W{50 + 3 + row}'] = row_name[1]
            worksheet[f'W{50 + 3 + row}'].border = plotting.default_border
            plotting.merge_cells(
                worksheet[f'W{50 + 3 + row}'],
                2,
                1
            )
            worksheet[f'Y{50 + 3 + row}'] = row_name[2]
            worksheet[f'Y{50 + 3 + row}'].border = plotting.default_border
            plotting.merge_cells(
                worksheet[f'Y{50 + 3 + row}'],
                2,
                1
            )
            worksheet[f'AA{50 + 3 + row}'] = ldos_top_20_in_2.iloc[row].values[
                0]
            worksheet[f'AA{50 + 3 + row}'].border = plotting.default_border