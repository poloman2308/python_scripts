from openpyxl.styles import Font, numbers
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from utils.web_app import \
    plotting_openpyxl as plotting
from utils.web_app import \
    Utils as Utils


class ExecutiveSummary:
    """ Build Get Healthy: Executive Summary sheet.
        Provide stats related to assets /w actions:
        [None, Refresh HW, Upgrade SW, Add Coverage]
    """

    def __init__(self, workbook: Workbook, devices):
        self.workbook = workbook
        self.devices = devices
        self.refresh_hw = self.get_refresh_hw()
        self.upgrade_sw = self.get_upgrade_sw()
        self.none = self.get_none()

    def get_refresh_hw(self):
        """ Get Devices with Recommended Action: Refresh Hardware
        """
        return self.devices.loc[(self.devices['Recommended Action'] == 'Refresh Hardware')]

    def get_upgrade_sw(self):
        """ Get Devices with Recommended Action: Upgrade Software
        """
        return self.devices.loc[(self.devices['Recommended Action'] == 'Upgrade Software')]

    def get_none(self):
        """ Get Devices with Recommended Action: None
        """
        return self.devices.loc[(self.devices['Recommended Action'] == 'No Action Required')]

    def create(self):
        """Create Executive Summary sheet."""
        plot = plotting
        utils = Utils

        worksheet: Worksheet = self.workbook.create_sheet('Executive Summary')

        worksheet.column_dimensions["A"].width = 2.6
        worksheet.column_dimensions["G"].width = 2.6
        worksheet.column_dimensions["M"].width = 2.6
        worksheet.column_dimensions["S"].width = 2.6
        worksheet.column_dimensions["Y"].width = 2.6

        worksheet['B2'] = 'Total Devices'
        worksheet['B2'].style = plotting.style_options
        plotting.merge_cells(
            worksheet['B2'],
            11,
            3
        )
        worksheet['N2'] = len(self.devices)

        worksheet['N2'].style = plotting.style_options
        plotting.merge_cells(
            worksheet['N2'],
            11,
            3
        )
        worksheet['N2'].number_format = numbers.BUILTIN_FORMATS[3]

        worksheet['B30'] = 'Recommended Action'
        worksheet['B30'].style = plotting.style_options
        plotting.merge_cells(
            worksheet['B30'],
            5,
            2
        )
        worksheet['H30'] = 'Device Count'
        worksheet['H30'].style = plotting.style_options
        plotting.merge_cells(
            worksheet['H30'],
            5,
            2
        )
        worksheet['N30'] = 'Device Percentage'
        worksheet['N30'].style = plotting.style_options
        plotting.merge_cells(
            worksheet['N30'],
            5,
            2
        )
        worksheet['T30'] = 'Explanation'
        worksheet['T30'].style = plotting.style_options
        plotting.merge_cells(
            worksheet['T30'],
            5,
            2
        )

        worksheet['B33'] = 'Refresh Hardware'
        worksheet['B33'].style = plotting.style_big_options
        plotting.merge_cells(
            worksheet['B33'],
            5,
            4
        )
        worksheet['B38'] = 'Upgrade Software'
        worksheet['B38'].style = plotting.style_big_options
        plotting.merge_cells(
            worksheet['B38'],
            5,
            4
        )
        worksheet['B43'] = 'No Action Required'
        worksheet['B43'].style = plotting.style_big_options
        plotting.merge_cells(
            worksheet['B43'],
            5,
            4
        )

        worksheet['H33'] = len(self.refresh_hw)
        worksheet['H33'].style = plotting.style_big_options_red
        plotting.merge_cells(
            worksheet['H33'],
            5,
            4
        )
        worksheet['H33'].number_format = numbers.BUILTIN_FORMATS[3]
        worksheet['H38'] = len(self.upgrade_sw)
        worksheet['H38'].style = plotting.style_big_options_yellow
        plotting.merge_cells(
            worksheet['H38'],
            5,
            4
        )
        worksheet['H38'].number_format = numbers.BUILTIN_FORMATS[3]
        worksheet['H43'] = len(self.none)
        worksheet['H43'].style = plotting.style_big_options_green
        plotting.merge_cells(
            worksheet['H43'],
            5,
            4
        )
        worksheet['H43'].number_format = numbers.BUILTIN_FORMATS[3]


        worksheet['N33'] = round(len(self.refresh_hw)/len(self.devices), 4)
        worksheet['N33'].style = plotting.style_big_options_red
        plotting.merge_cells(
            worksheet['N33'],
            5,
            4
        )
        worksheet['N33'].number_format = numbers.FORMAT_PERCENTAGE_00
        worksheet['N38'] = round(len(self.upgrade_sw)/len(self.devices), 4)
        worksheet['N38'].style = plotting.style_big_options_yellow
        plotting.merge_cells(
            worksheet['N38'],
            5,
            4
        )
        worksheet['N38'].number_format = numbers.FORMAT_PERCENTAGE_00
        worksheet['N43'] = round(len(self.none)/len(self.devices), 4)
        worksheet['N43'].style = plotting.style_big_options_green
        plotting.merge_cells(
            worksheet['N43'],
            5,
            4
        )
        worksheet['N43'].number_format = numbers.FORMAT_PERCENTAGE_00

        worksheet['T33'] = 'HW LDoS is =< 12 months'
        worksheet['T33'].style = plotting.style_small_options_txt
        plotting.merge_cells(
            worksheet['T33'],
            5,
            4
        )
        worksheet['T38'] = 'HW LDoS is > 12 months\nSW EoSWM is =< 12 months'
        worksheet['T38'].style = plotting.style_small_options_txt
        plotting.merge_cells(
            worksheet['T38'],
            5,
            4
        )
        worksheet['T43'] = 'HW LDoS is > 12 months\nSW EoSWM is > 12 months'
        worksheet['T43'].style = plotting.style_small_options_txt
        plotting.merge_cells(
            worksheet['T43'],
            5,
            4
        )

        pivot = self.devices.pivot_table(index=['Recommended Action'], values=['Serial Number'], aggfunc='count').sort_index()
        plot.create_pie(worksheet, pivot, 'pie', 'orange', 'Recommended Action',
                        '', '', 8, 6, True, 5.5, 6.5, summary=True)

        worksheet['B49'] = 'Upgrade Software'
        worksheet['B49'].style = plotting.style_big_options_yellow
        plotting.merge_cells(
            worksheet['B49'],
            11,
            3
        )

        upgrade_sw = self.upgrade_sw.pivot_table(index=['Shortened Product Family'], values=['Serial Number'], aggfunc='count').sort_values(('Serial Number'), ascending=False).head(20)
        upgrade_sw = upgrade_sw.rename(columns={"Serial Number": "Components"})
        max = upgrade_sw.iloc[0].values[0]
        min = upgrade_sw.iloc[len(upgrade_sw)-1].values[0]
        for row in range(0, len(upgrade_sw)):
            font_size = 10+(20*((upgrade_sw.iloc[row].values[0]-min)/(max-min)))
            worksheet[f'B{49+3+row}'] = upgrade_sw.iloc[row].name
            worksheet[f'B{49 + 3 + row}'].font = Font(size=font_size)
            worksheet[f'B{49 + 3 + row}'].border = plotting.default_border
            plotting.merge_cells(
                worksheet[f'B{49+3+row}'],
                9,
                1
            )
            worksheet[f'K{49+3+row}'] = upgrade_sw.iloc[row].values[0]
            worksheet[f'K{49 + 3 + row}'].font = Font(size=font_size)
            worksheet[f'K{49 + 3 + row}'].border = plotting.default_border
            plotting.merge_cells(
                worksheet[f'K{49+3+row}'],
                2,
                1
            )

        worksheet['N49'] = 'Refresh Hardware'
        plotting.merge_cells(
            worksheet['N49'],
            11,
            3
        )
        worksheet['N49'].style = plotting.style_big_options_red
        refresh_hw = self.refresh_hw.pivot_table(index=['Shortened Product Family'], values=['Serial Number'], aggfunc='count').sort_values(('Serial Number'), ascending=False).head(20)
        refresh_hw = refresh_hw.rename(columns={"Serial Number": "Components"})
        max = refresh_hw.iloc[0].values[0]
        min = refresh_hw.iloc[len(refresh_hw)-1].values[0]
        for row in range(0, len(refresh_hw)):
            font_size = 10+(20*((refresh_hw.iloc[row].values[0]-min)/(max-min)))
            worksheet[f'N{49+3+row}'] = refresh_hw.iloc[row].name
            worksheet[f'N{49 + 3 + row}'].font = Font(size=font_size)
            worksheet[f'N{49 + 3 + row}'].border = plotting.default_border
            plotting.merge_cells(
                worksheet[f'N{49+3+row}'],
                9,
                1
            )
            worksheet[f'W{49+3+row}'] = refresh_hw.iloc[row].values[0]
            worksheet[f'W{49 + 3 + row}'].font = Font(size=font_size)
            worksheet[f'W{49 + 3 + row}'].border = plotting.default_border
            plotting.merge_cells(
                worksheet[f'W{49+3+row}'],
                2,
                1
            )

        worksheet.sheet_view.showGridLines = False
        